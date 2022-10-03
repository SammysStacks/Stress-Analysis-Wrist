

# Basic Modules
import os
import sys
import numpy as np
import pandas as pd
# Read/Write to Excel
import csv
import pyexcel
import openpyxl as xl
# Openpyxl Styles
from openpyxl.styles import Alignment
from openpyxl.styles import Font



class handlingExcelFormat:        
        
    def convertToXLSX(self, inputExcelFile):
        """
        Converts .xls Files to .xlsx Files That OpenPyxl Can Read
        If the File is Already a .xlsx Files, Do Nothing
        If the File is Neither a .xls Nor .xlsx, it Exits the Program
        """
        # Check That the Current Extension is .xls or .xlsx
        _, extension = os.path.splitext(inputExcelFile)
        # If the Extension is .xlsx, the File is Ready; Do Nothing
        if extension == '.xlsx':
            return inputExcelFile
        # If the Extension is Not .xls/.xlsx, Then the Data is in the Wrong Format; Exit Program
        if extension not in ['.xls', '.xlsx']:
            print("Cannot Convert File to .xlsx")
            sys.exit()
        
        # Create Output File Directory to Save Data ONLY If None Exists
        newExcelFolder = os.path.dirname(inputExcelFile) + "/Excel Files/"
        os.makedirs(newExcelFolder, exist_ok = True)
        
        # Convert '.xls' to '.xlsx'
        filename = os.path.basename(inputExcelFile)
        newExcelFile = newExcelFolder + filename + "x"
        pyexcel.save_as(file_name = inputExcelFile, dest_file_name = newExcelFile, logfile=open(os.devnull, 'w'))
        
        # Save New Excel name
        return newExcelFile
    
    def txt2csv(self, txtFile, csvFile, csvDelimiter = ",", overwriteCSV = False):
        # Check to see if csv conversion alreayd happened
        if not os.path.isfile(csvFile) or overwriteCSV:
            with open(txtFile, "r") as inputData:
                in_reader = csv.reader(inputData, delimiter = csvDelimiter)
                with open(csvFile, 'w', newline='') as out_csv:
                    out_writer = csv.writer(out_csv)
                    for row in in_reader:
                        out_writer.writerow(row)
    
    def convertToExcel(self, inputFile, excelFile, excelDelimiter = ",", overwriteXL = False, testSheetNum = 0):
        # If the File is Not Already Converted: Convert the CSV to XLSX
        if not os.path.isfile(excelFile) or overwriteXL:
            if excelDelimiter == "fixedWidth":
                df = pd.read_fwf(inputFile)
                df.drop(index=0, inplace=True) # drop the underlines
                df.to_excel(excelFile, index=False)
                # Load the Data from the Excel File
                xlWorkbook = xl.load_workbook(excelFile, data_only=True, read_only=True)
                xlWorksheet = xlWorkbook.worksheets[testSheetNum]
            else:
                # Make Excel WorkBook
                xlWorkbook = xl.Workbook()
                xlWorksheet = xlWorkbook.active
                # Write the Data from the CSV File to the Excel WorkBook
                with open(inputFile, "r") as inputData:
                    inReader = csv.reader(inputData, delimiter = excelDelimiter)
                    with open(excelFile, 'w+', newline=''):
                        for row in inReader:
                            xlWorksheet.append(row)    
                # Save as New Excel File
                xlWorkbook.save(excelFile)
        # Else Load the Data from the Excel File
        else:
            # Load the Data from the Excel File
            xlWorkbook = xl.load_workbook(excelFile, data_only=True, read_only=True)
            xlWorksheet = xlWorkbook.worksheets[testSheetNum]
        
        # Return Excel Sheet
        return xlWorkbook, xlWorksheet

    def addExcelAesthetics(self, WB_worksheet):
        # Center the Data in the Cells
        align = Alignment(horizontal='center',vertical='center',wrap_text=True)        
        for column_cells in WB_worksheet.columns:
            length = max(len(str(cell.value) if cell.value else "") for cell in column_cells)
            WB_worksheet.column_dimensions[xl.utils.get_column_letter(column_cells[0].column)].width = length
            
            for cell in column_cells:
                cell.alignment = align
        # Header Style
        for cell in WB_worksheet["1:1"]:
            cell.font = Font(color='00FF0000', italic=True, bold=True)
        
        return WB_worksheet

class dataProcessing(handlingExcelFormat):

    def saveResults(self, featureList, featureLabels, saveDataFolder, saveExcelName, sheetName = "Pulse Features", overwriteSave = True, dontSaveIfExcelExists = False):
        print("Saving the Data")
        # Create Output File Directory to Save Data: If None Exists
        os.makedirs(saveDataFolder, exist_ok=True)
        
        # Path to File to Save
        excelFile = saveDataFolder + saveExcelName
        
        # If You Want to Overwrite the Excel, Remove the File First (Quicker)
        if overwriteSave and os.path.isfile(excelFile):
            print("\tDeleting Old Excel Workbook")
            os.remove(excelFile) 
        
        # If the File is Not Present: Create The Excel File
        if not os.path.isfile(excelFile):
            print("\tSaving the Data as New Excel Workbook")
            # Make Excel WorkBook
            WB = xl.Workbook()
            WB_worksheet = WB.active 
            WB_worksheet.title = sheetName
        elif not dontSaveIfExcelExists:
            print("\tExcel File Already Exists. Adding New Sheet to File")
            WB = xl.load_workbook(excelFile, read_only=False)
            WB_worksheet = WB.create_sheet(sheetName)
        else:
            print("\tNot Saving Any Data as the Excel File Already Exists")
            return None
    
        # Parameters for Worksheet
        header = featureLabels    # Header Text     
        maxAddToExcelSheet = 1048500  # Max Rows in a Worksheet
        
        # Save Data to Worksheet
        for firstIndexInList in range(0, len(featureList), maxAddToExcelSheet):
            # Add the Header to the Worksheet
            WB_worksheet.append(header)
            
            # Add the Features to the Worksheet
            for featureInd in range(firstIndexInList, min(firstIndexInList+maxAddToExcelSheet, len(featureList))):
                WB_worksheet.append(list(featureList[featureInd]))
        
            # Add Excel Aesthetics
            WB_worksheet = self.addExcelAesthetics(WB_worksheet)  
            
            # Add Sheet
            WB_worksheet = WB.create_sheet(sheetName)
        
        WB.remove(WB_worksheet)
        # Save as New Excel File
        WB.save(excelFile)
        WB.close()

    def getSavedFeatures(self, featureExcelFile):
        # Check if File Exists
        if not os.path.exists(featureExcelFile):
            print("The following Input File Does Not Exist:", featureExcelFile)
            sys.exit()

        print("Extracting Features from the Excel File:", featureExcelFile)
        # Load Data from the Excel File
        WB = xl.load_workbook(featureExcelFile, data_only=True, read_only=True)
        WB_worksheets = WB.worksheets
        ExcelSheet = WB_worksheets[-1]
        
        # If Header Exists, Skip Until You Find the Data
        for row in ExcelSheet.rows:
            cellA = row[0]
            if type(cellA.value) in [float, int]:
                dataStartRow = cellA.row
                break
        
        features = []
        # Loop Through the Excel Worksheet to collect all the data
        for row in ExcelSheet.iter_rows(min_col=1, min_row=dataStartRow, max_row=ExcelSheet.max_row):
            # SafeGaurd: If User Edits the Document to Create Empty Rows, Stop Reading in Data
            if row[0].value == None:
                break
            
            newFeatures = []
            for cell in row:
                if cell.value == None:
                    break
                newFeatures.append(cell.value)
            features.append(newFeatures)
            
        # Finished Data Collection: Close Workbook and Return Data to User
        WB.close()
        return np.array(features)
    
    def extractFeatureNames(self, featureLabelFile, prependedString, appendToName = ''):
        """ Extract the Feature Names from a txt File """
        # Check if File Exists
        if not os.path.exists(featureLabelFile):
            print("The following Input File Does Not Exist:", featureLabelFile)
            sys.exit()

        # Get the Data
        fullText = ''
        with open(featureLabelFile, "r", newline='\n') as inputData:
            inReader = csv.reader(inputData)
            for row in inReader:
                for featureString in row:
                    if featureString[0] != "#":
                        fullText += featureString + ","
        
        possibleFeatures = fullText.split(prependedString)
        # Extract the Features
        featureList = []
        for feature in possibleFeatures:
            feature = feature.split("[")[-1]
            feature = feature.split("]")[0]
            feature = feature.replace(" ", "")
            feature = feature.replace("\n", "")
            
            if len(feature) != 0:
                feature = feature.split(",")
                featureList.extend(feature)
                
        featureListFull = []
        for feature in featureList:
            featureListFull.append(feature + appendToName)
        
        return featureListFull

class processPulseData(dataProcessing):
    
    def getData(self, pulseExcelFile, testSheetNum = 0):
        """
        Extracts Pulse Data from Excel Document (.xlsx). Data can be in any
        worksheet which the user can specify using 'testSheetNum' (0-indexed).
        In the Worksheet:
            Time Data must be in Column 'A' (x-Axis)
            Capacitance Data must be in Column 'B' (y-Axis)
        If No Data is present in one cell of a row, it will be read in as zero.
        --------------------------------------------------------------------------
        Input Variable Definitions:
            pulseExcelFile: The Path to the Excel File Containing the Pulse Data
            testSheetNum: An Integer Representing the Excel Worksheet (0-indexed) Order.
        --------------------------------------------------------------------------
        """
        # Check if File Exists
        if not os.path.exists(pulseExcelFile):
            print("The following Input File Does Not Exist:", pulseExcelFile)
            sys.exit()
        # Convert to Exel if .xls Format; If .xlsx, Do Nothing; If Other, Exit Program
        pulseExcelFile = self.convertToXLSX(pulseExcelFile)

        print("Extracting Data from the Excel File:", pulseExcelFile)
        # Load Data from the Excel File
        WB = xl.load_workbook(pulseExcelFile, data_only=True, read_only=True)
        WB_worksheets = WB.worksheets
        ExcelSheet = WB_worksheets[testSheetNum]
        
        # If Header Exists, Skip Until You Find the Data
        for row in ExcelSheet.rows:
            cellA = row[0]
            if type(cellA.value) in [float, int]:
                dataStartRow = cellA.row
                break
        
        data = dict(time=[], Capacitance=[])
        # Loop Through the Excel Worksheet to collect all the data
        for pointNum, [colA, colB] in enumerate(ExcelSheet.iter_rows(min_col=1, min_row=dataStartRow, max_col=2, max_row=ExcelSheet.max_row)):
            # Get Cell Values for First 4 Channels: Represents the Voltage for Each Channel;
            Time = colA.value; Capacitance = colB.value;
            
            # SafeGaurd: If User Edits the Document to Create Empty Rows, Stop Reading in Data
            if Time == None or Capacitance == None:
                break
            
            # Add Data to Dictionary
            data["time"].append(Time)
            data["Capacitance"].append(Capacitance)
             
        # Finished Data Collection: Close Workbook and Return Data to User
        print("Done Data Collecting"); WB.close()
        return np.array(data["time"]), np.array(data["Capacitance"])
    
    def saveFilteredData(self, time, signalData, filteredData, saveDataFolder, saveExcelName, sheetName = "Pulse Data"):
        print("Saving the Data")
        # Create Output File Directory to Save Data: If None Exists
        os.makedirs(saveDataFolder, exist_ok=True)
        
        # Path to File to Save
        excelFile = saveDataFolder + saveExcelName
        
        # If the File is Not Present: Create The Excel File
        if not os.path.isfile(excelFile):
            print("\tSaving the Data as New Excel Workbook")
            # Make Excel WorkBook
            WB = xl.Workbook()
            WB_worksheet = WB.active 
            WB_worksheet.title = sheetName
        else:
            print("Excel File Already Exists. Adding New Sheet to File")
            WB = xl.load_workbook(excelFile, read_only=False)
            WB_worksheet = WB.create_sheet(sheetName)
        
        # Label First Row
        header = ["Time", "Data", "Filtered Data"]
        WB_worksheet.append(header)
        
        # Save Data to Worksheet
        for pulseInd in range(len(time)):
            row = [time[pulseInd], signalData[pulseInd], filteredData[pulseInd]]
            WB_worksheet.append(row)
        
        # Add Excel Aesthetics
        WB_worksheet = self.addExcelAesthetics(WB_worksheet)  
        
        # Save as New Excel File
        WB.save(excelFile)
        WB.close()
    
class processTemperatureData(dataProcessing):
    
    def extractTemperatureData(self, ExcelSheet, startDataCol = 1, endDataCol = 2):
        # If Header Exists, Skip Until You Find the Data
        for row in ExcelSheet.rows:
            cellA = row[0]
            if type(cellA.value) in [int, float]:
                dataStartRow = cellA.row + 1
                break
        
        # Loop Through the Excel Worksheet to collect all the data
        timePoints = []; gsrData = [];
        for dataRow in ExcelSheet.iter_rows(min_col=startDataCol, min_row=dataStartRow-1, max_col=endDataCol, max_row=ExcelSheet.max_row):
            # Stop Collecting Data When there is No More
            if dataRow[0].value == None:
                break
            
            # Get Data
            gsrData.append(float(dataRow[1].value))
            timePoints.append(float(dataRow[0].value))
        
        return timePoints, gsrData

    def getData(self, inputFile, testSheetNum = 0):
        """
        Extracts Pulse Data from Excel Document (.xlsx). Data can be in any
        worksheet which the user can specify using 'testSheetNum' (0-indexed).
        In the Worksheet:
            Time Data must be in Column 'A' (x-Axis)
            Capacitance Data must be in Column 'B' (y-Axis)
        If No Data is present in one cell of a row, it will be read in as zero.
        --------------------------------------------------------------------------
        Input Variable Definitions:
            tempFile: The Path to the Excel/TXT/CSV File Containing the GSR Data
            testSheetNum: An Integer Representing the Excel Worksheet (0-indexed) Order.
        --------------------------------------------------------------------------
        """
        # Check if File Exists
        if not os.path.exists(inputFile):
            print("The following Input File Does Not Exist:", inputFile)
            sys.exit()
            
        # Convert to TXT and CSV Files to XLSX
        if inputFile.endswith(".txt") or inputFile.endswith(".csv"):
            # Extract Filename Information
            oldFileExtension = os.path.basename(inputFile)
            filename = os.path.splitext(oldFileExtension)[0]
            newFilePath = os.path.dirname(inputFile) + "/Excel Files/"
            # Make Output Folder Directory if Not Already Created
            os.makedirs(newFilePath, exist_ok = True)

            # Convert CSV or TXT to XLSX
            excelFile = newFilePath + filename + ".xlsx"
            xlWorkbook, chiWorksheet = self.convertToExcel(inputFile, excelFile, excelDelimiter = ",", overwriteXL = False, testSheetNum = testSheetNum)
        # If the File is Already an Excel File, Just Load the File
        elif inputFile.endswith(".xlsx"):
            # Load the GSR Data from the Excel File
            xlWorkbook = xl.load_workbook(inputFile, data_only=True, read_only=True)
            chiWorksheet = xlWorkbook.worksheets[testSheetNum]
        else:
            print("The Following File is Neither CSV, TXT, Nor XLSX:", inputFile)
        print("Extracting Data from the Excel File:", inputFile)
        
        # Extract Time and Current Data from the File
        timePoints, temperature = self.extractTemperatureData(chiWorksheet)

        
        xlWorkbook.close()
        # Finished Data Collection: Close Workbook and Return Data to User
        print("Done Collecting GSR Data");
        return np.array(timePoints), np.array(temperature)

class processGSRData(dataProcessing):
    
    def extractCHIData_CurrentTime(self, chiWorksheet):
        
        # -------------------------------------------------------------------#
        # ----------------------- Extract Run Info --------------------------#
        
        # Get Time and Current Data from Excel as Well as CHI Labeled Peaks
        findStart = True
        timePoints = []; currentPoints = []
        peakTimesCHI = []; peakCurrentsCHI = []; peakAmplitudesCHI = []
        # Loop Through the Info Section and Extract the Needxed Run Info from Excel
        rowGenerator = chiWorksheet.rows
        for cell in rowGenerator:
            # Get Cell Value
            cellVal = cell[0].value
            
            # Extract Improtant Information from the File
            if findStart:
                # If Nothing in Cell, Continue
                if cellVal == None:
                    continue
                # If Time Peak Found by CHI, Store the Value
                elif cellVal.startswith("tp = "):
                    peakTimeVal = float(cellVal.split(" = ")[-1][:-1])
                    peakTimesCHI.append(peakTimeVal)
                # If Time Peak Found by CHI, Store the Value
                elif cellVal.startswith("ip = "):
                    peakCurrentVal = float(cellVal.split(" = ")[-1][:-1])
                    peakCurrentsCHI.append(peakCurrentVal)
                # If Amplitude Peak Found by CHI, Store the Value
                elif cellVal.startswith("Ap = "):
                    peakAmplitudeVal = float(cellVal.split(" = ")[-1][:-1])
                    peakAmplitudesCHI.append(peakAmplitudeVal)
                # If Current/Time Titles are Present, the Data is Starting Soon
                elif cellVal == "Time/sec":
                    next(rowGenerator) # Skip Over Empty Cell After Title
                    findStart = False
            # Extract the Data from the File
            else:
                # Break out of Loop if no More Data (edge effect if someone edits excel)
                if cell[0].value == None:
                    break
                
                # Keep Track  the Time and Current Data points
                timePoints.append(float(cell[0].value))
                currentPoints.append(float(cell[1].value))
        
        # Return Time and Current
        currentPoints = np.array(currentPoints)
        timePoints = np.array(timePoints)
        return timePoints, currentPoints
    
    def extractGSRData(self, ExcelSheet, startDataCol = 1, endDataCol = 2):
        # If Header Exists, Skip Until You Find the Data
        for row in ExcelSheet.rows:
            cellA = row[0]
            if type(cellA.value) in [int, float]:
                dataStartRow = cellA.row + 1
                break
        
        # Loop Through the Excel Worksheet to collect all the data
        timePoints = []; gsrData = [];
        for dataRow in ExcelSheet.iter_rows(min_col=startDataCol, min_row=dataStartRow-1, max_col=endDataCol, max_row=ExcelSheet.max_row):
            # Stop Collecting Data When there is No More
            if dataRow[0].value == None:
                break
            
            # Get Data
            gsrData.append(float(dataRow[1].value))
            timePoints.append(float(dataRow[0].value))
        
        return timePoints, gsrData

    def getData(self, inputFile, testSheetNum = 0, method = "useCHI"):
        """
        Extracts Pulse Data from Excel Document (.xlsx). Data can be in any
        worksheet which the user can specify using 'testSheetNum' (0-indexed).
        In the Worksheet:
            Time Data must be in Column 'A' (x-Axis)
            Capacitance Data must be in Column 'B' (y-Axis)
        If No Data is present in one cell of a row, it will be read in as zero.
        --------------------------------------------------------------------------
        Input Variable Definitions:
            inputFile: The Path to the Excel/TXT/CSV File Containing the GSR Data
            testSheetNum: An Integer Representing the Excel Worksheet (0-indexed) Order.
        --------------------------------------------------------------------------
        """
        # Check if File Exists
        if not os.path.exists(inputFile):
            print("The following Input File Does Not Exist:", inputFile)
            sys.exit()
            
        # Convert to TXT and CSV Files to XLSX
        if inputFile.endswith(".txt") or inputFile.endswith(".csv"):
            # Extract Filename Information
            oldFileExtension = os.path.basename(inputFile)
            filename = os.path.splitext(oldFileExtension)[0]
            newFilePath = os.path.dirname(inputFile) + "/Excel Files/"
            # Make Output Folder Directory if Not Already Created
            os.makedirs(newFilePath, exist_ok = True)

            # Convert CSV or TXT to XLSX
            excelFile = newFilePath + filename + ".xlsx"
            xlWorkbook, chiWorksheet = self.convertToExcel(inputFile, excelFile, excelDelimiter = ",", overwriteXL = False, testSheetNum = testSheetNum)
        # If the File is Already an Excel File, Just Load the File
        elif inputFile.endswith(".xlsx"):
            # Load the GSR Data from the Excel File
            xlWorkbook = xl.load_workbook(inputFile, data_only=True, read_only=True)
            chiWorksheet = xlWorkbook.worksheets[testSheetNum]
        else:
            print("The Following File is Neither CSV, TXT, Nor XLSX:", inputFile)
        print("Extracting Data from the Excel File:", inputFile)
        
        # Extract Time and Current Data from the File
        if method == 'useCHI':
            timePoints, currentPoints = self.extractCHIData_CurrentTime(chiWorksheet)
        elif method == 'processed':
            timePoints, currentPoints = self.extractGSRData(chiWorksheet)
        else:
            exit("No Extract Method for GSR Found")
        
        xlWorkbook.close()
        # Finished Data Collection: Close Workbook and Return Data to User
        print("Done Collecting GSR Data");
        return np.array(timePoints), np.array(currentPoints)
    
    def saveFilteredData(self, timeGSR, currentGS, saveDataFolder, saveExcelName, sheetName = "Galvanic Skin Response Data"):
        print("Saving the Data")
        # Create Output File Directory to Save Data: If Not Already Created
        os.makedirs(saveDataFolder, exist_ok=True)
        
        # Create Path to Save the Excel File
        excelFile = saveDataFolder + saveExcelName
        
        # If the File is Not Present: Create it
        if not os.path.isfile(excelFile):
            # Make Excel WorkBook
            WB = xl.Workbook()
            WB_worksheet = WB.active 
            WB_worksheet.title = sheetName
        else:
            print("Excel File Already Exists. Adding New Sheet to File")
            WB = xl.load_workbook(excelFile)
            WB_worksheet = WB.create_sheet(sheetName)
        
        # Label First Row
        header = ["Time (Seconds)", "Current (uAmps)"]
        WB_worksheet.append(header)
        
        # Save Data to Worksheet
        for dataPoint in range(len(timeGSR)):
            timePoint = timeGSR[dataPoint]
            currentPoint = currentGS[dataPoint]
            
            # Write the Data to Excel
            WB_worksheet.append([timePoint, currentPoint])
        
        # Add Excel Aesthetics
        WB_worksheet = self.addExcelAesthetics(WB_worksheet)    
            
        # Save as New Excel File
        WB.save(excelFile)
        WB.close()


class processChemicalData(dataProcessing):
    
    def extractChemicalData(self, ExcelSheet, startDataCol = 1, endDataCol = 4):
        # If Header Exists, Skip Until You Find the Data
        for row in ExcelSheet.rows:
            cellA = row[0]
            if type(cellA.value) in [int, float]:
                dataStartRow = cellA.row + 1
                break
        
        # Loop Through the Excel Worksheet to collect all the data
        timePoints = []; glucose = []; lactate = []; uricAcid = []
        for dataRow in ExcelSheet.iter_rows(min_col=startDataCol, min_row=dataStartRow-1, max_col=endDataCol, max_row=ExcelSheet.max_row):
            # Stop Collecting Data When there is No More
            if dataRow[0].value == None:
                break
            
            # Get Data
            glucoseData = dataRow[1].value
            lactateData = dataRow[2].value
            uricAcidData = dataRow[3].value

            # Get Cell Values
            timePoints.append(dataRow[0].value)
            if glucoseData != None:
                glucose.append(float(glucoseData))
            if lactateData != None:
                lactate.append(float(lactateData))
            if uricAcidData != None:
                uricAcid.append(float(uricAcidData))
        
        return timePoints, np.array([np.array(glucose), np.array(lactate), np.array(uricAcid)])
            
    def getData(self, chemicalFile, testSheetNum = 0):
        """
        Extracts Pulse Data from Excel Document (.xlsx). Data can be in any
        worksheet which the user can specify using 'testSheetNum' (0-indexed).
        In the Worksheet:
            Time Data must be in Column 'A' (x-Axis)
            Capacitance Data must be in Column 'B' (y-Axis)
        If No Data is present in one cell of a row, it will be read in as zero.
        --------------------------------------------------------------------------
        Input Variable Definitions:
            chemicalFile: The Path to the Excel/TXT/CSV File Containing the Chemical Data
            testSheetNum: An Integer Representing the Excel Worksheet (0-indexed) Order.
        --------------------------------------------------------------------------
        """
        # Check if File Exists
        if not os.path.exists(chemicalFile):
            print("The following Input File Does Not Exist:", chemicalFile)
            sys.exit()
            
        # Convert to TXT and CSV Files to XLSX
        if chemicalFile.endswith(".txt") or chemicalFile.endswith(".csv"):
            # Extract Filename Information
            oldFileExtension = os.path.basename(chemicalFile)
            filename = os.path.splitext(oldFileExtension)[0]
            newFilePath = os.path.dirname(chemicalFile) + "/Excel Files/"
            # Make Output Folder Directory if Not Already Created
            os.makedirs(newFilePath, exist_ok = True)

            # Convert CSV or TXT to XLSX
            excelFile = newFilePath + filename + ".xlsx"
            xlWorkbook, chemicalWorksheet = self.convertToExcel(chemicalFile, excelFile, excelDelimiter = ",", overwriteXL = False, testSheetNum = testSheetNum)
        # If the File is Already an Excel File, Just Load the File
        elif chemicalFile.endswith(".xlsx"):
            # Load the GSR Data from the Excel File
            xlWorkbook = xl.load_workbook(chemicalFile, data_only=True, read_only=True)
            chemicalWorksheet = xlWorkbook.worksheets[testSheetNum]
        else:
            print("The Following File is Neither CSV, TXT, Nor XLSX:", chemicalFile)
        print("Extracting Data from the Excel File:", chemicalFile)
        
        # Extract Time and Current Data from the File
        timePoints, chemicalData = self.extractChemicalData(chemicalWorksheet)
        
        xlWorkbook.close()
        # Finished Data Collection: Close Workbook and Return Data to User
        print("Done Collecting Chemical Data");
        return np.array(timePoints), np.array(chemicalData)

class processMLData(dataProcessing):
    
    def saveFeatureComparison(self, dataMatrix, rowHeaders, colHeaders, saveDataFolder, saveExcelName, sheetName = "Feature Comparison", saveFirstSheet = False):
        print("Saving the Data")
        # Create Output File Directory to Save Data: If Not Already Created
        os.makedirs(saveDataFolder, exist_ok=True)
        
        # Create Path to Save the Excel File
        excelFile = saveDataFolder + saveExcelName
        
        # If the File is Not Present: Create it
        if not os.path.isfile(excelFile):
            # Make Excel WorkBook
            WB = xl.Workbook()
            WB_worksheet = WB.active 
            WB_worksheet.title = sheetName
        else:
            print("Excel File Already Exists. Adding New Sheet to File")
            WB = xl.load_workbook(excelFile)
            WB_worksheet = WB.create_sheet(sheetName)
        

        maxAddToExcelSheet = 1048500  # Max Rows in a Worksheet
        # Save Data to Worksheet
        for firstIndexInList in range(0, len(dataMatrix), maxAddToExcelSheet):
            # Label First Row
            WB_worksheet.append(colHeaders)
            
            # Add data to the Worksheet
            for rowInd in range(firstIndexInList, min(firstIndexInList+maxAddToExcelSheet, len(dataMatrix))):
                dataRow = []
                
                if rowInd < len(rowHeaders):
                    rowHeader = rowHeaders[rowInd]
                    dataRow.append(rowHeader)
                
                dataRow.extend(dataMatrix[rowInd])
                dataRow[0] = float(dataRow[0])
                # Write the Data to Excel
                WB_worksheet.append(dataRow)
        
            # Add Excel Aesthetics
            WB_worksheet = self.addExcelAesthetics(WB_worksheet)  
            # Add Sheet
            WB_worksheet = WB.create_sheet(sheetName)
            
            if saveFirstSheet:
                break

        WB.remove(WB_worksheet)
        # Save as New Excel File
        WB.save(excelFile)
        WB.close()
    
    def getData(self, MLFile, signalData = [], signalLabels = [], testSheetNum = 0, startCollectionCol = 2):
        """
        Extracts Pulse Data from Excel Document (.xlsx). Data can be in any
        worksheet which the user can specify using 'testSheetNum' (0-indexed).
        
        If No Data is present in one cell of a row, it will be read in as zero.
        --------------------------------------------------------------------------
        Input Variable Definitions:
            MLFile: The Path to the Excel File Containing the Compiled ML Data
            testSheetNum: An Integer Representing the Excel Worksheet (0-indexed) Order.
            startCollectionCol: The Column Index (1-Indexed) Of the Start Time of the Pulse
        --------------------------------------------------------------------------
        """
        # Check if File Exists
        if not os.path.exists(MLFile):
            print("The following Input File Does Not Exist:", MLFile)
            sys.exit()
        # Convert to Exel if .xls Format; If .xlsx, Do Nothing; If Other, Exit Program
        if MLFile.endswith(('.xlsx', '.xls', '.csv', 'txt')):
            MLFile = self.convertToXLSX(MLFile)
        else:
            return signalData, signalLabels, []

        # Load Data from the Excel File
        WB = xl.load_workbook(MLFile, data_only=True, read_only=True)
        WB_worksheets = WB.worksheets
        ExcelSheet = WB_worksheets[testSheetNum]
        
        # If Header Exists, Skip Until You Find the Data
        headerTitles = []
        for row in ExcelSheet.rows:
            cellA = row[0]
            # Find When the Data Starts: When First Number Appears
            if type(cellA.value) in [float, int]:
                dataStartRow = cellA.row
                break
            # Find the Header: Last String Before the Data
            elif type(cellA) == str:
                headerTitles = ["Pulse Time"]
                for col in row[startCollectionCol+1:-1]:
                    if col.value != None:
                        headerTitles.append(col.value)
        
        # Loop Through the Excel Worksheet to collect all the data
        for pointNum, row in enumerate(ExcelSheet.iter_rows(min_col=startCollectionCol, max_col=ExcelSheet.max_col, min_row=dataStartRow, max_row=ExcelSheet.max_row)):
            # SafeGaurd: If User Edits the Document to Create Empty Rows, Stop Reading in Data
            if len(row) == 0 or row[0].value == None:
                break
            
            # The First Two Columns MUST be the Pulse Times
            signalData.append([row[1].value - row[0].value])
            # Get Cell Values for the Data
            for cellCol in row[2:-1]:
                if cellCol.value != None:
                    signalData[-1].append(cellCol.value)
        # Add Signal Label (Should be at the End of the Data)
        signalLabels.append(row[-1].value)
             
        # Finished Data Collection: Close Workbook and Return Data to User
        return signalData, signalLabels, headerTitles
  
    
    